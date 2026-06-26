# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Carregar leads limpos
with open('leads_limpos.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Leads de prioridade (sem site próprio válido)
leads_prioridade = data.get('leads_prioridade', [])

# Leads rejeitados na limpeza
leads_rejeitados = data.get('rejeitados_limpeza', [])

# Combinar todos os leads
todos_leads = leads_prioridade + leads_rejeitados

# Filtrar leads sem site E com WhatsApp (apenas pet shop e odontologia do JSON original)
leads_originais = []
for lead in todos_leads:
    red_flags = lead.get('red_flags', [])
    whatsapp = lead.get('whatsapp', '')
    segmento = lead.get('segmento', '')
    
    if ('sem_site_proprio_valido' in red_flags and 
        whatsapp and whatsapp.strip() and 
        segmento in ['odontologia', 'pet_shop']):
        leads_originais.append(lead)

# Novos leads fornecidos pelo usuário
novos_leads = [
    {
        "nome": "Moretto Clínica Odontológica",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 3384-3244",
        "whatsapp": "5511994449755",
        "whatsapp_link": "https://api.whatsapp.com/send?phone=5511994449755",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Moretto%20Cl%C3%ADnica%20Odontol%C3%B3gica&query_place_id=ChIJAY6ZAzRfzpQRXSo3BV8ualM"
    },
    {
        "nome": "Dentz Clínica Odontológica",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 94604-6000",
        "whatsapp": "5511946046000",
        "whatsapp_link": "https://api.whatsapp.com/send?phone=5511946046000&text=Ol%C3%A1,%20quero%20agendar%20uma%20avalia%C3%A7%C3%A3o!",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Dentz%20Cl%C3%ADnica%20Odontol%C3%B3gica&query_place_id=ChIJa3Oiki5bzpQRFSt7T7FSUCc"
    },
    {
        "nome": "Clínica Jardim Novo Horizonte",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 5938-8967",
        "whatsapp": "551159388967",
        "whatsapp_link": "https://wa.me/551159388967",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Cl%C3%ADnica%20Jardim%20Novo%20Horizonte&query_place_id=ChIJ6146NbpJzpQRJKtfbqK3OCM"
    },
    {
        "nome": "Dentista Oriél Odontologia Modernna",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 98500-5666",
        "whatsapp": "5511985005666",
        "whatsapp_link": "https://wa.me/5511985005666?text=Ol%C3%A1%2C%20vim%20pelo%20google%20e%20quero%20saber%20mais",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Dentista%20Ori%C3%A9l%20Odontologia%20Modernna%20I%20Clinica%20Odontol%C3%B3gica%20em%20Cidade%20Dutra%20Interlagos&query_place_id=ChIJ18HtAaZPzpQRvL2MXaXf2x0"
    },
    {
        "nome": "Clínica Odontológica Hoffmann Pires",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 91585-8959",
        "whatsapp": "5511915858959",
        "whatsapp_link": "https://wa.me/5511915858959",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Cl%C3%ADnica%20Odontol%C3%B3gica%20Hoffmann%20Pires&query_place_id=ChIJ0cxopGJRzpQRiwJGzkk4zf8"
    },
    {
        "nome": "DentGold Clínica Odontológica - Jabaquara",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 2667-2823",
        "whatsapp": "5511989762024",
        "whatsapp_link": "https://api.whatsapp.com/send?phone=5511989762024&text=Ol%C3%A1!",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=DentGold%20Cl%C3%ADnica%20Odontol%C3%B3gica%20-%20Jabaquara&query_place_id=ChIJufWFzdtazpQRmp1awT-3OY0"
    },
    {
        "nome": "CQB ODONTOLOGIA",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 94072-3340",
        "whatsapp": "5511940723340",
        "whatsapp_link": "https://api.whatsapp.com/send/?phone=5511940723340&text=Ol%C3%A1%2C+Gostaria+de+agendar+uma+consulta%21&type=phone_number&app_absent=0",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=CQB%20ODONTOLOGIA&query_place_id=ChIJASnriCRRzpQRnlwoh5wOqmM"
    },
    {
        "nome": "Dra Priscila Galhasso - Clínica Odontologica Galhasso",
        "segmento": "odontologia",
        "cidade": "São Paulo",
        "estado": "SP",
        "telefone": "+55 11 93470-7944",
        "whatsapp": "5511934707944",
        "whatsapp_link": "https://api.whatsapp.com/send?phone=5511934707944&text=Ol%C3%A1,%20Venho%20do%20google%20e%20estou%20interessado(a)%20em%20seus%20servi%C3%A7os",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Dra%20Priscila%20Galhasso%20-%20Cl%C3%ADnica%20Odontologica%20Galhasso&query_place_id=ChIJHyXN1r_3zpQR5cY095dD7bQ"
    },
    {
        "nome": "Clínica Odontológica DNTBRAS",
        "segmento": "odontologia",
        "cidade": "Juiz de Fora",
        "estado": "MG",
        "telefone": "+55 32 99882-9667",
        "whatsapp": "5532998628476",
        "whatsapp_link": "https://wa.me/5532998628476",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Cl%C3%ADnica%20Odontol%C3%B3gica%20DNTBRAS&query_place_id=ChIJqfimO4OdmAARkpbCe5VWyq4"
    },
    {
        "nome": "Odontologia Jobim",
        "segmento": "odontologia",
        "cidade": "Juiz de Fora",
        "estado": "MG",
        "telefone": "+55 32 99858-4597",
        "whatsapp": "5532998584597",
        "whatsapp_link": "https://wa.me/5532998584597",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Odontologia%20Jobim&query_place_id=ChIJAZ50rW6bmAARkLGf4KrEuCY"
    },
    {
        "nome": "Khárisma Odontologia - Reabilitação Oral com Dr. Gabriel David",
        "segmento": "odontologia",
        "cidade": "Juiz de Fora",
        "estado": "MG",
        "telefone": "+55 32 3015-2913",
        "whatsapp": "5532998211868",
        "whatsapp_link": "https://wa.me/5532998211868?text=Ol%C3%A1%2C%20vi%20voc%C3%AAs%20no%20Google%20e%20quero%20saber%20mais%20sobre%20implantes%20dent%C3%A1rios.",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Kh%C3%A1risma%20Odontologia%20%7C%20Reabilita%C3%A7%C3%A3o%20Oral%20com%20Dr.%20Gabriel%20David&query_place_id=ChIJNzb6LnibmAARv_UJkrrWrBU"
    },
    {
        "nome": "Clínica Médica e Odontológica Santa Rosa",
        "segmento": "odontologia",
        "cidade": "Montes Claros",
        "estado": "MG",
        "telefone": "+55 38 99903-5439",
        "whatsapp": "5538999035439",
        "whatsapp_link": "https://wa.me/5538999035439",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Cl%C3%ADnica%20M%C3%A9dica%20e%20Odontol%C3%B3gica%20%7C%20Santa%20Rosa&query_place_id=ChIJuw5afQBVqwARIlPwwmMYkcw"
    },
    {
        "nome": "OdontoMontes",
        "segmento": "odontologia",
        "cidade": "Montes Claros",
        "estado": "MG",
        "telefone": "+55 38 3015-1002",
        "whatsapp": "553899606242",
        "whatsapp_link": "http://api.whatsapp.com/send/?phone=553899606242&text=Ol%C3%A1,+vi+seu+an%C3%BAncio+no+google+e+tenho+interesse!&type=phone_number&app_absent=0",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=OdontoMontes&query_place_id=ChIJgQFvIkhVqwAR1MEaEa0mfYE"
    },
    {
        "nome": "Consultório Odontológico Dra Lorena Rocha Unidade 2",
        "segmento": "odontologia",
        "cidade": "Montes Claros",
        "estado": "MG",
        "telefone": "+55 38 98405-6942",
        "whatsapp": "5538984056942",
        "whatsapp_link": "https://wa.me/5538984056942",
        "maps_link": "https://www.google.com/maps/search/?api=1&query=Consult%C3%B3rio%20Odontol%C3%B3gico%20Dra%20Lorena%20Rocha%20Unidade%202&query_place_id=ChIJKZBxjhpVqwAR53oB4UP3LPw"
    },
    {
        "nome": "Clínica Caiafa - Dra Luísa Caiafa",
        "segmento": "odontologia",
        "cidade": "Juiz de Fora",
        "estado": "MG",
        "telefone": "+55 32 3026-4165",
        "whatsapp": "",
        "whatsapp_link": "",
        "maps_link": ""
    }
]

# Combinar todos os leads
todos_leads_final = leads_originais + novos_leads

# Cores
verde_escuro = "1F4E3D"
verde_medio = "5F6F52"
verde_claro = "E8F5E9"
dourado = "B87D5E"
branco = "FFFFFF"
cinza_claro = "F5F5F5"

# Estilos
fonte_titulo = Font(name='Calibri', bold=True, size=16, color=branco)
fonte_cabecalho = Font(name='Calibri', bold=True, size=11, color=branco)
fill_cabecalho = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
fill_par = PatternFill(start_color=cinza_claro, end_color=cinza_claro, fill_type="solid")
fill_impar = PatternFill(start_color=branco, end_color=branco, fill_type="solid")
fonte_dados = Font(name='Calibri', size=10)
fonte_link = Font(name='Calibri', size=10, color="0563C1", underline='single')
alinhamento_cabecalho = Alignment(horizontal='center', vertical='center', wrap_text=True)
alinhamento_dados = Alignment(vertical='center', wrap_text=True)
borda_fina = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Criar workbook
wb = Workbook()

# ==========================================
# ABA 1: RESUMO
# ==========================================
ws_resumo = wb.active
ws_resumo.title = "Resumo"

# Título
ws_resumo.merge_cells('A1:D1')
cell_titulo = ws_resumo['A1']
cell_titulo.value = "LEADS SEM SITE - ODONTOLOGIA & PET SHOP"
cell_titulo.font = fonte_titulo
cell_titulo.fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
ws_resumo.row_dimensions[1].height = 40

# Totais
total = len(todos_leads_final)
odontologia = sum(1 for l in todos_leads_final if l.get('segmento') == 'odontologia')
pet_shop = sum(1 for l in todos_leads_final if l.get('segmento') == 'pet_shop')

# Card Total
ws_resumo.merge_cells('A3:B3')
ws_resumo['A3'].value = "TOTAL DE LEADS"
ws_resumo['A3'].font = Font(name='Calibri', size=11, color="666666")
ws_resumo['A3'].alignment = Alignment(horizontal='center')

ws_resumo.merge_cells('A4:B4')
ws_resumo['A4'].value = total
ws_resumo['A4'].font = Font(name='Calibri', bold=True, size=24, color=verde_escuro)
ws_resumo['A4'].alignment = Alignment(horizontal='center')
ws_resumo['A4'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")

# Card Odontologia
ws_resumo['C3'].value = "ODONTOLOGIA"
ws_resumo['C3'].font = Font(name='Calibri', size=11, color="666666")
ws_resumo['C3'].alignment = Alignment(horizontal='center')

ws_resumo['C4'].value = odontologia
ws_resumo['C4'].font = Font(name='Calibri', bold=True, size=24, color=dourado)
ws_resumo['C4'].alignment = Alignment(horizontal='center')
ws_resumo['C4'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")

# Card Pet Shop
ws_resumo['D3'].value = "PET SHOP"
ws_resumo['D3'].font = Font(name='Calibri', size=11, color="666666")
ws_resumo['D3'].alignment = Alignment(horizontal='center')

ws_resumo['D4'].value = pet_shop
ws_resumo['D4'].font = Font(name='Calibri', bold=True, size=24, color=verde_medio)
ws_resumo['D4'].alignment = Alignment(horizontal='center')
ws_resumo['D4'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")

# Ajustar largura
ws_resumo.column_dimensions['A'].width = 20
ws_resumo.column_dimensions['B'].width = 20
ws_resumo.column_dimensions['C'].width = 20
ws_resumo.column_dimensions['D'].width = 20

# ==========================================
# ABA 2: LEADS
# ==========================================
ws_leads = wb.create_sheet(title="Leads")

# Cabeçalhos
headers = ["#", "NOME", "SEGMENTO", "CIDADE", "ESTADO", "WHATSAPP", "LINK WHATSAPP", "TELEFONE", "LINK MAPS"]

for col, header in enumerate(headers, 1):
    cell = ws_leads.cell(row=1, column=col, value=header)
    cell.font = fonte_cabecalho
    cell.fill = fill_cabecalho
    cell.alignment = alinhamento_cabecalho
    cell.border = borda_fina

ws_leads.row_dimensions[1].height = 30

# Larguras
larguras = [5, 35, 15, 25, 8, 20, 45, 20, 55]
for i, largura in enumerate(larguras, 1):
    ws_leads.column_dimensions[get_column_letter(i)].width = largura

# Dados
for i, lead in enumerate(todos_leads_final, 1):
    row = i + 1
    fill_atual = fill_par if i % 2 == 0 else fill_impar
    
    nome = lead.get('nome', 'N/A')
    segmento = "Odontologia" if lead.get('segmento') == 'odontologia' else "Pet Shop"
    cidade = lead.get('cidade', 'N/A')
    estado = lead.get('estado', 'N/A')
    whatsapp = lead.get('whatsapp', 'N/A')
    telefone = lead.get('telefone', 'N/A')
    
    # Links
    whatsapp_link = lead.get('whatsapp_link', '')
    if not whatsapp_link and whatsapp:
        whatsapp_link = f"https://wa.me/{whatsapp}"
    
    maps_link = lead.get('maps_link', '')
    if not maps_link:
        place_id = lead.get('place_id', 'N/A')
        if place_id and place_id != 'N/A':
            maps_link = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
    
    dados = [i, nome, segmento, cidade, estado, whatsapp, whatsapp_link, telefone, maps_link]
    
    for col, valor in enumerate(dados, 1):
        cell = ws_leads.cell(row=row, column=col, value=valor)
        cell.fill = fill_atual
        cell.border = borda_fina
        cell.alignment = alinhamento_dados
        
        if col in [7, 9]:
            cell.font = fonte_link
            if valor:
                cell.hyperlink = valor
        else:
            cell.font = fonte_dados
    
    ws_leads.row_dimensions[row].height = 25

ws_leads.freeze_panes = 'A2'
ws_leads.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(todos_leads_final) + 1}"

# ==========================================
# ABA 3: POR SEGMENTO
# ==========================================
ws_segmento = wb.create_sheet(title="Por Segmento")

# Título
ws_segmento.merge_cells('A1:F1')
ws_segmento['A1'].value = f"LEADS POR SEGMENTO ({total} TOTAL)"
ws_segmento['A1'].font = fonte_titulo
ws_segmento['A1'].fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
ws_segmento['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_segmento.row_dimensions[1].height = 35

# Cabeçalhos
headers_segmento = ["#", "NOME", "CIDADE", "WHATSAPP", "LINK WHATSAPP", "LINK MAPS"]
for col, header in enumerate(headers_segmento, 1):
    cell = ws_segmento.cell(row=3, column=col, value=header)
    cell.font = fonte_cabecalho
    cell.fill = fill_cabecalho
    cell.alignment = alinhamento_cabecalho
    cell.border = borda_fina

larguras_segmento = [5, 35, 25, 20, 45, 55]
for i, largura in enumerate(larguras_segmento, 1):
    ws_segmento.column_dimensions[get_column_letter(i)].width = largura

# Separar por segmento
odontologia_leads = [l for l in todos_leads_final if l.get('segmento') == 'odontologia']
pet_shop_leads = [l for l in todos_leads_final if l.get('segmento') == 'pet_shop']

# Odontologia
row_atual = 4
ws_segmento.merge_cells(f'A{row_atual}:F{row_atual}')
ws_segmento[f'A{row_atual}'].value = f"ODONTOLOGIA ({len(odontologia_leads)} leads)"
ws_segmento[f'A{row_atual}'].font = Font(name='Calibri', bold=True, size=12, color=dourado)
ws_segmento[f'A{row_atual}'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")
ws_segmento.row_dimensions[row_atual].height = 25
row_atual += 1

for i, lead in enumerate(odontologia_leads, 1):
    fill_atual = fill_par if i % 2 == 0 else fill_impar
    whatsapp = lead.get('whatsapp', 'N/A')
    whatsapp_link = lead.get('whatsapp_link', '')
    if not whatsapp_link and whatsapp:
        whatsapp_link = f"https://wa.me/{whatsapp}"
    maps_link = lead.get('maps_link', '')
    if not maps_link:
        place_id = lead.get('place_id', 'N/A')
        if place_id and place_id != 'N/A':
            maps_link = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
    
    dados = [i, lead.get('nome', 'N/A'), f"{lead.get('cidade', 'N/A')} - {lead.get('estado', 'N/A')}", 
             whatsapp, whatsapp_link, maps_link]
    
    for col, valor in enumerate(dados, 1):
        cell = ws_segmento.cell(row=row_atual, column=col, value=valor)
        cell.fill = fill_atual
        cell.border = borda_fina
        cell.alignment = alinhamento_dados
        if col in [5, 6]:
            cell.font = fonte_link
            if valor:
                cell.hyperlink = valor
        else:
            cell.font = fonte_dados
    
    ws_segmento.row_dimensions[row_atual].height = 25
    row_atual += 1

# Espaço
row_atual += 1

# Pet Shop
ws_segmento.merge_cells(f'A{row_atual}:F{row_atual}')
ws_segmento[f'A{row_atual}'].value = f"PET SHOP ({len(pet_shop_leads)} leads)"
ws_segmento[f'A{row_atual}'].font = Font(name='Calibri', bold=True, size=12, color=verde_medio)
ws_segmento[f'A{row_atual}'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")
ws_segmento.row_dimensions[row_atual].height = 25
row_atual += 1

for i, lead in enumerate(pet_shop_leads, 1):
    fill_atual = fill_par if i % 2 == 0 else fill_impar
    whatsapp = lead.get('whatsapp', 'N/A')
    whatsapp_link = lead.get('whatsapp_link', '')
    if not whatsapp_link and whatsapp:
        whatsapp_link = f"https://wa.me/{whatsapp}"
    maps_link = lead.get('maps_link', '')
    if not maps_link:
        place_id = lead.get('place_id', 'N/A')
        if place_id and place_id != 'N/A':
            maps_link = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
    
    dados = [i, lead.get('nome', 'N/A'), f"{lead.get('cidade', 'N/A')} - {lead.get('estado', 'N/A')}", 
             whatsapp, whatsapp_link, maps_link]
    
    for col, valor in enumerate(dados, 1):
        cell = ws_segmento.cell(row=row_atual, column=col, value=valor)
        cell.fill = fill_atual
        cell.border = borda_fina
        cell.alignment = alinhamento_dados
        if col in [5, 6]:
            cell.font = fonte_link
            if valor:
                cell.hyperlink = valor
        else:
            cell.font = fonte_dados
    
    ws_segmento.row_dimensions[row_atual].height = 25
    row_atual += 1

# Salvar
wb.save('leads_sem_site_whatsapp.xlsx')
print("Excel atualizado com sucesso!")
print(f"Total de leads: {total}")
print(f"  - Odontologia: {odontologia}")
print(f"  - Pet Shop: {pet_shop}")
print(f"  - Leads originais: {len(leads_originais)}")
print(f"  - Novos leads adicionados: {len(novos_leads)}")