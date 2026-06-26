# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Carregar leads limpos
with open('leads_limpos.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Leads de prioridade (sem site próprio válido)
leads_prioridade = data.get('leads_prioridade', [])

# Leads rejeitados na limpeza
leads_rejeitados = data.get('rejeitados_limpeza', [])

# Combinar todos os leads
todos_leads = leads_prioridade + leads_rejeitados

# Filtrar leads sem site E com WhatsApp
leads_sem_site_com_whatsapp = []
for lead in todos_leads:
    red_flags = lead.get('red_flags', [])
    whatsapp = lead.get('whatsapp', '')
    segmento = lead.get('segmento', '')
    
    if ('sem_site_proprio_valido' in red_flags and 
        whatsapp and whatsapp.strip() and 
        segmento in ['odontologia', 'pet_shop']):
        leads_sem_site_com_whatsapp.append(lead)

# Ordenar por segmento
leads_sem_site_com_whatsapp.sort(key=lambda x: x.get('segmento', ''))

# Criar workbook
wb = Workbook()

# ==========================================
# ABA 1: RESUMO
# ==========================================
ws_resumo = wb.active
ws_resumo.title = "Resumo"

# Cores
verde_escuro = "1F4E3D"
verde_medio = "5F6F52"
verde_claro = "E8F5E9"
dourado = "B87D5E"
branco = "FFFFFF"
cinza_claro = "F5F5F5"

# Estilos de fonte
fonte_titulo = Font(name='Calibri', bold=True, size=16, color=branco)
fonte_subtitulo = Font(name='Calibri', bold=True, size=12, color=verde_escuro)
fonte_numero = Font(name='Calibri', bold=True, size=24, color=verde_escuro)
fonte_label = Font(name='Calibri', size=11, color="666666")

# Preencher resumo
ws_resumo.merge_cells('A1:D1')
cell_titulo = ws_resumo['A1']
cell_titulo.value = "LEADS SEM SITE - ODONTOLOGIA & PET SHOP"
cell_titulo.font = fonte_titulo
cell_titulo.fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
ws_resumo.row_dimensions[1].height = 40

# Dados do resumo
total = len(leads_sem_site_com_whatsapp)
odontologia = sum(1 for l in leads_sem_site_com_whatsapp if l.get('segmento') == 'odontologia')
pet_shop = sum(1 for l in leads_sem_site_com_whatsapp if l.get('segmento') == 'pet_shop')

# Card Total
ws_resumo.merge_cells('A3:B3')
ws_resumo['A3'].value = "TOTAL DE LEADS"
ws_resumo['A3'].font = fonte_label
ws_resumo['A3'].alignment = Alignment(horizontal='center')

ws_resumo.merge_cells('A4:B4')
ws_resumo['A4'].value = total
ws_resumo['A4'].font = fonte_numero
ws_resumo['A4'].alignment = Alignment(horizontal='center')
ws_resumo['A4'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")

# Card Odontologia
ws_resumo.merge_cells('C3:C3')
ws_resumo['C3'].value = "ODONTOLOGIA"
ws_resumo['C3'].font = fonte_label
ws_resumo['C3'].alignment = Alignment(horizontal='center')

ws_resumo.merge_cells('C4:C4')
ws_resumo['C4'].value = odontologia
ws_resumo['C4'].font = Font(name='Calibri', bold=True, size=24, color=dourado)
ws_resumo['C4'].alignment = Alignment(horizontal='center')
ws_resumo['C4'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")

# Card Pet Shop
ws_resumo['D3'].value = "PET SHOP"
ws_resumo['D3'].font = fonte_label
ws_resumo['D3'].alignment = Alignment(horizontal='center')

ws_resumo['D4'].value = pet_shop
ws_resumo['D4'].font = Font(name='Calibri', bold=True, size=24, color=verde_medio)
ws_resumo['D4'].alignment = Alignment(horizontal='center')
ws_resumo['D4'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")

# Ajustar largura das colunas do resumo
ws_resumo.column_dimensions['A'].width = 20
ws_resumo.column_dimensions['B'].width = 20
ws_resumo.column_dimensions['C'].width = 20
ws_resumo.column_dimensions['D'].width = 20

# ==========================================
# ABA 2: LEADS
# ==========================================
ws_leads = wb.create_sheet(title="Leads")

# Cabeçalhos
headers = ["#", "NOME", "SEGMENTO", "CIDADE", "ESTADO", "WHATSAPP", "LINK WHATSAPP", "TELEFONE", "LINK MAPS"]

# Estilo do cabeçalho
fonte_cabecalho = Font(name='Calibri', bold=True, size=11, color=branco)
fill_cabecalho = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
alinhamento_cabecalho = Alignment(horizontal='center', vertical='center', wrap_text=True)
borda_fina = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Escrever cabeçalhos
for col, header in enumerate(headers, 1):
    cell = ws_leads.cell(row=1, column=col, value=header)
    cell.font = fonte_cabecalho
    cell.fill = fill_cabecalho
    cell.alignment = alinhamento_cabecalho
    cell.border = borda_fina

# Ajustar largura das colunas
larguras = [5, 30, 15, 25, 8, 20, 40, 20, 50]
for i, largura in enumerate(larguras, 1):
    ws_leads.column_dimensions[get_column_letter(i)].width = largura

# Altura da linha de cabeçalho
ws_leads.row_dimensions[1].height = 30

# Preencher dados
fill_par = PatternFill(start_color=cinza_claro, end_color=cinza_claro, fill_type="solid")
fill_impar = PatternFill(start_color=branco, end_color=branco, fill_type="solid")
fonte_dados = Font(name='Calibri', size=10)
fonte_link = Font(name='Calibri', size=10, color="0563C1", underline='single')
alinhamento_dados = Alignment(vertical='center', wrap_text=True)

for i, lead in enumerate(leads_sem_site_com_whatsapp, 1):
    row = i + 1
    fill_atual = fill_par if i % 2 == 0 else fill_impar
    
    # Dados
    nome = lead.get('nome', 'N/A')
    segmento = "Odontologia" if lead.get('segmento') == 'odontologia' else "Pet Shop"
    cidade = lead.get('cidade', 'N/A')
    estado = lead.get('estado', 'N/A')
    whatsapp = lead.get('whatsapp', 'N/A')
    telefone = lead.get('telefone', 'N/A')
    place_id = lead.get('place_id', 'N/A')
    url_original = lead.get('url_original_maps', '')
    
    # Links
    whatsapp_link = f"https://wa.me/{whatsapp}" if whatsapp else "N/A"
    maps_link = f"https://www.google.com/maps/place/?q=place_id:{place_id}" if place_id and place_id != 'N/A' else url_original
    
    # Escrever dados
    dados = [i, nome, segmento, cidade, estado, whatsapp, whatsapp_link, telefone, maps_link]
    
    for col, valor in enumerate(dados, 1):
        cell = ws_leads.cell(row=row, column=col, value=valor)
        cell.fill = fill_atual
        cell.border = borda_fina
        cell.alignment = alinhamento_dados
        
        # Aplicar fonte de link para colunas de links
        if col in [7, 9]:  # Colunas de link
            cell.font = fonte_link
            cell.hyperlink = valor
        else:
            cell.font = fonte_dados
    
    # Altura da linha
    ws_leads.row_dimensions[row].height = 25

# Congelar primeira linha
ws_leads.freeze_panes = 'A2'

# Filtro automático
ws_leads.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads_sem_site_com_whatsapp) + 1}"

# ==========================================
# ABA 3: POR SEGMENTO
# ==========================================
ws_segmento = wb.create_sheet(title="Por Segmento")

# Título
ws_segmento.merge_cells('A1:F1')
ws_segmento['A1'].value = "LEADS POR SEGMENTO"
ws_segmento['A1'].font = fonte_titulo
ws_segmento['A1'].fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type="solid")
ws_segmento['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_segmento.row_dimensions[1].height = 35

# Cabeçalhos
headers_segmento = ["#", "NOME", "CIDADE", "WHATSAPP", "LINK WHATSAPP", "LINK MAPS"]
for col, header in enumerate(headers_segmento, 1):
    cell = ws_segmento.cell(row=3, column=col, value=header)
    cell.font = fonte_cabecalho
    cell.fill = fill_cabecalho
    cell.alignment = alinhamento_cabecalho
    cell.border = borda_fina

# Larguras
larguras_segmento = [5, 30, 25, 20, 40, 50]
for i, largura in enumerate(larguras_segmento, 1):
    ws_segmento.column_dimensions[get_column_letter(i)].width = largura

# Separar por segmento
odontologia_leads = [l for l in leads_sem_site_com_whatsapp if l.get('segmento') == 'odontologia']
pet_shop_leads = [l for l in leads_sem_site_com_whatsapp if l.get('segmento') == 'pet_shop']

# Seção Odontologia
row_atual = 4
ws_segmento.merge_cells(f'A{row_atual}:F{row_atual}')
ws_segmento[f'A{row_atual}'].value = f"ODONTOLOGIA ({len(odontologia_leads)} leads)"
ws_segmento[f'A{row_atual}'].font = Font(name='Calibri', bold=True, size=12, color=dourado)
ws_segmento[f'A{row_atual}'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")
ws_segmento.row_dimensions[row_atual].height = 25
row_atual += 1

for i, lead in enumerate(odontologia_leads, 1):
    fill_atual = fill_par if i % 2 == 0 else fill_impar
    whatsapp = lead.get('whatsapp', 'N/A')
    place_id = lead.get('place_id', 'N/A')
    url_original = lead.get('url_original_maps', '')
    whatsapp_link = f"https://wa.me/{whatsapp}" if whatsapp else "N/A"
    maps_link = f"https://www.google.com/maps/place/?q=place_id:{place_id}" if place_id and place_id != 'N/A' else url_original
    
    dados = [i, lead.get('nome', 'N/A'), f"{lead.get('cidade', 'N/A')} - {lead.get('estado', 'N/A')}", 
             whatsapp, whatsapp_link, maps_link]
    
    for col, valor in enumerate(dados, 1):
        cell = ws_segmento.cell(row=row_atual, column=col, value=valor)
        cell.fill = fill_atual
        cell.border = borda_fina
        cell.alignment = alinhamento_dados
        if col in [5, 6]:
            cell.font = fonte_link
            cell.hyperlink = valor
        else:
            cell.font = fonte_dados
    
    ws_segmento.row_dimensions[row_atual].height = 25
    row_atual += 1

# Espaço
row_atual += 1

# Seção Pet Shop
ws_segmento.merge_cells(f'A{row_atual}:F{row_atual}')
ws_segmento[f'A{row_atual}'].value = f"PET SHOP ({len(pet_shop_leads)} leads)"
ws_segmento[f'A{row_atual}'].font = Font(name='Calibri', bold=True, size=12, color=verde_medio)
ws_segmento[f'A{row_atual}'].fill = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")
ws_segmento.row_dimensions[row_atual].height = 25
row_atual += 1

for i, lead in enumerate(pet_shop_leads, 1):
    fill_atual = fill_par if i % 2 == 0 else fill_impar
    whatsapp = lead.get('whatsapp', 'N/A')
    place_id = lead.get('place_id', 'N/A')
    url_original = lead.get('url_original_maps', '')
    whatsapp_link = f"https://wa.me/{whatsapp}" if whatsapp else "N/A"
    maps_link = f"https://www.google.com/maps/place/?q=place_id:{place_id}" if place_id and place_id != 'N/A' else url_original
    
    dados = [i, lead.get('nome', 'N/A'), f"{lead.get('cidade', 'N/A')} - {lead.get('estado', 'N/A')}", 
             whatsapp, whatsapp_link, maps_link]
    
    for col, valor in enumerate(dados, 1):
        cell = ws_segmento.cell(row=row_atual, column=col, value=valor)
        cell.fill = fill_atual
        cell.border = borda_fina
        cell.alignment = alinhamento_dados
        if col in [5, 6]:
            cell.font = fonte_link
            cell.hyperlink = valor
        else:
            cell.font = fonte_dados
    
    ws_segmento.row_dimensions[row_atual].height = 25
    row_atual += 1

# Salvar
output_path = "leads_sem_site_whatsapp.xlsx"
wb.save(output_path)
print(f"Excel gerado com sucesso: {output_path}")
print(f"Total de leads: {total}")
print(f"  - Odontologia: {odontologia}")
print(f"  - Pet Shop: {pet_shop}")